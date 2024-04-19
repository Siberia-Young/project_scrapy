from scrapy_selenium import SeleniumRequest
import scrapy
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
from selenium.webdriver.common.proxy import Proxy, ProxyType
import shutil
import json
import time
import pygetwindow as gw

from project_scrapy.modules.tools.tip import tip

from project_scrapy.items import Item

class JDDetailSpider(scrapy.Spider):
    name = "jd_detail"
    def start_requests(self):
        tip(self.name)
        # 读取parameter.json的参数
        with open('parameter.json', encoding='utf-8') as file:
            parameter = json.load(file)
            self.graphicwords = parameter['graphicwords']
        url = "https://www.jd.com"
        yield SeleniumRequest(url=url, callback=self.parse, wait_time=2)
        
    def parse(self, response):
        file_path = '../data/jd/merge/merge_2_3_8_9.xlsx'
        workbook = load_workbook(file_path)
        sheet = workbook.active
        driver = response.webdriver
        try:
            copy_file_path = file_path.replace('.xlsx','(副本).xlsx')
            shutil.copy(file_path, copy_file_path)
            temp_workbook = load_workbook(copy_file_path)
            temp_sheet = temp_workbook.active
            for row in range(2, temp_sheet.max_row+1):
                temp_sheet.cell(row=row, column=16, value='')
            temp_workbook.save(copy_file_path)
        except:
            print(f'\n出错')

        start_time = time.time()

        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        time.sleep(2)

        try:
            count = 0
            end = False
            while count != 0 or end == False:
                count = 0
                if end == True:
                    current = 0
                    print(f'\n新一轮处理')
                for row in range(start_row, end_row + 1):
                    current+=1
                    res = (total - current) / (current / ((time.time() - start_time) / 60))
                    print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")

                    # 如果品牌字段不为空，且该行被标记应该被删除，这跳过不处理
                    temp = sheet.cell(row=row, column=9).value
                    delete = sheet.cell(row=row, column=15).value
                    if (temp != None) or (delete != None and delete == 'delete'):
                        continue

                    goods_link = sheet.cell(row=row, column=10).value
                    driver.get(goods_link)

                    try:
                        # 缓慢下拉页面
                        scroll_height = driver.execute_script("return document.body.scrollHeight;")
                        scroll_count = 0
                        current_height = 0
                        scroll_speed = 500  # 每次下拉的距离
                        while scroll_count < 3:
                            driver.execute_script(f"window.scrollTo(0, {current_height});")
                            current_height += scroll_speed
                            scroll_count += 1
                            time.sleep(0.1)  # 等待一段时间，模拟缓慢下拉的效果
                        time.sleep(0.2)
                    except:
                        print('下拉获取页面信息时发生错误')

                    tempHTML = driver.execute_script("return document.documentElement.outerHTML")
                    tempSoup = BeautifulSoup(tempHTML, "html.parser")

                    # 遇到验证页面或者登录页面人工解决
                    verify = tempSoup.select('div.verifyBtn')
                    login = tempSoup.select('div.login-btn')
                    while len(verify) != 0 or len(login) != 0:
                        firefox_window = gw.getWindowsWithTitle("Mozilla Firefox")[0]
                        firefox_window.minimize()
                        firefox_window.maximize()
                        firefox_window.activate()
                        time.sleep(7)
                        html = driver.execute_script(
                            "return document.documentElement.outerHTML")
                        soup = BeautifulSoup(html, "html.parser") 
                        verify = soup.select('div.verifyBtn')
                        login = tempSoup.select('div.login-btn')

                    elements = tempSoup.select('div.hxm_hide_page')
                    if len(elements) == 0:
                        elements = tempSoup.select('div.itemover-tip')
                    if len(elements) == 0:
                        elements = tempSoup.select('div.logo_extend')
                    if len(elements) == 0:
                        try:
                            goods_brand_element = tempSoup.find_all('ul',id='parameter-brand')
                            if len(goods_brand_element) != 0:
                                goods_brand = goods_brand_element[0].select('li')[0].text
                                if goods_brand.startswith('品牌：'):
                                    goods_brand = goods_brand.replace('品牌：', '').replace('\n', '').replace('\r', '').replace(' ', '')
                                else:
                                    goods_brand = "暂无"
                            else:
                                goods_brand = "暂无"
                            
                            sheet.cell(row=row, column=9, value=goods_brand)
                            
                            choose = tempSoup.select('div.li.p-choose:not(.hide)')
                            choose_text_list = []
                            for item in choose:
                                choose_list = item.select('div.dd div a')
                                for item1 in choose_list:
                                    choose_text_list.append(item1.text.strip())
                            sheet.cell(row=row, column=17, value='\n'.join(choose_text_list))

                            parameter = tempSoup.select('div.p-parameter ul.p-parameter-list')
                            parameter_text_list = []
                            for item in parameter:
                                parameter_list = item.select('li')
                                for item1 in parameter_list:
                                    parameter_text_list.append(item1.text.strip())
                            sheet.cell(row=row, column=18, value='\n'.join(parameter_text_list))

                            detail_img = tempSoup.select('div.spec-items ul.lh li img')
                            if len(detail_img) != 0:
                                img_src = 'https:' + detail_img[0].get('src')[:-5]
                                base_src = sheet.cell(row=row, column=7).value
                                src1 = '/'.join(img_src.split('/')[-2:]).split('.')[0]
                                src2 = '/'.join(base_src.split('/')[-2:]).split('.')[0]
                                if src1 != src2:
                                    sheet.cell(row=row, column=19, value='different')
                        except:
                            workbook.save(file_path)
                            print('与现有浏览器连接断开')
                    else:
                        sheet.cell(row=row, column=15, value='delete')
                    count += 1
                end = True
        except Exception as e:
            print(e)
            print('主动中断')
        finally:
            # 保存文件
            workbook.save(file_path)
            print('与现有浏览器连接断开')
            end_time = time.time()
            duration = end_time - start_time
            print(f"爬虫耗时：{duration:.2f} 秒")
            print(f"目标数量：{total} 条")
            print(f"已获取数量：{current} 条")
            unit = current / (duration / 60)
            print(f"每分钟爬取数量：{unit:.2f} 条")

            item = Item()
            item['category'] = 'JD_process_3'
            item['end'] = True
            item['graphicwords'] = self.graphicwords
            yield item