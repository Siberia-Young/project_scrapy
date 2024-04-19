from scrapy_selenium import SeleniumRequest
from openpyxl import Workbook
from bs4 import BeautifulSoup
import urllib.parse
import scrapy
import datetime
import random
import json
import pygetwindow as gw
import time
import os
from openpyxl import load_workbook

from project_scrapy.modules.tools.tip import tip

from project_scrapy.items import Item

class JDBasicSpider(scrapy.Spider):
    name = "jd_basic"
    def start_requests(self):
        tip(self.name)
        self.min_delay = 3  # 最小延迟时间（单位：秒）
        self.max_delay = 5  # 最大延迟时间（单位：秒）

        # 读取parameter.json的参数
        with open('parameter.json', encoding='utf-8') as file:
            parameter = json.load(file)
            self.keywords = parameter['keywords']
            self.relatedwords = parameter['relatedwords']
            self.graphicwords = parameter['graphicwords']
            self.start_page = parameter['start_page']
            self.end_page = parameter['end_page']*2
        
        self.process_status = {}
        for keyword in self.keywords:
            self.process_status[keyword] = False
            
        url = "https://www.jd.com"
        yield SeleniumRequest(url=url, callback=self.parse, wait_time=2)
        
        
    def parse(self, response):
        driver = response.webdriver
        # 针对各个关键词构造url
        for keyword in self.keywords:
            [file_path, headers] = self.file_prepare(keyword)
            url = f"https://search.jd.com/Search?keyword={urllib.parse.quote(keyword)}&psort=4&page=1&s=1"
            driver.get(url)
            time.sleep(5)
            
            self.verify_monitor(driver)

            html = driver.execute_script("return document.documentElement.outerHTML")
            soup = BeautifulSoup(html, "html.parser") 
            
            start_page = self.start_page
            end_page = self.get_max_page(soup)
            total_num = 0
            record_num = 0
            try:
                for page in range(start_page, end_page+1):
                    [single_total_num, single_record_num, item_page] = self.parse_page(driver, page, headers, file_path, keyword)
                    total_num += single_total_num
                    record_num += single_record_num
                    if single_total_num != 0 and single_record_num == 0:
                        break
                    item = Item()
                    item['category'] = 'JD_spider'
                    item['filepath'] = file_path
                    item['keyword'] = keyword
                    item['page'] = page
                    item['goods'] = item_page
                    yield item
            except Exception as e:
                print(e)
                print('记录每页数据时发生错误')
            except KeyboardInterrupt:
                print('用户主动中断爬虫')
            finally:
                # 重命名文件
                new_file_path = file_path.replace('.xlsx',f'_({record_num} of {total_num}).xlsx')
                try:
                    os.rename(file_path, new_file_path)
                    print(f"已将文件 {file_path} 重命名为 {new_file_path}")
                except Exception as e:
                    print(e)
                    print(f"重命名文件 {file_path} 失败")
                finally:
                    print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")
                    self.process_status[keyword] = True
                    print(f'关键词爬取完毕：{keyword}')

                    item = Item()
                    item['category'] = 'JD_process_1'
                    item['end'] = True
                    item['keywords'] = self.keywords
                    item['relatedwords'] = self.relatedwords
                    item['graphicwords'] = self.graphicwords
                    for value in self.process_status.values():
                        if not value:
                            item['end'] = False
                    yield item

    def parse_page(self, driver, page, headers, file_path, keyword):
        # 加载之前创建的excel表
        workbook = load_workbook(file_path)
        sheet = workbook.active
        last_row = sheet.max_row

        # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
        total_num = 0
        record_num = 0

        url = f"https://search.jd.com/Search?keyword={urllib.parse.quote(keyword)}&psort=4&page={str(page)}&s=1"
        driver.get(url)

        # 设置适当的时间间隔以避免触发反爬虫
        time.sleep(random.uniform(self.min_delay, self.max_delay))

        self.verify_monitor(driver)
        
        # 通过bs4得到整个页面解析得来的单个商品元素构成的列表
        html = driver.execute_script("return document.documentElement.outerHTML")
        soup = BeautifulSoup(html, "html.parser") 
        elements = soup.select('li.gl-item')

        item_page = []
        try:
            # 逐个解析提取列表中商品信息的各部分数据
            for (index, element) in enumerate(elements):
                item_row = {}

                shop_elements = element.select('div.p-shop a.curr-shop.hd-shopname')
                goods_elements = element.select('div.p-img a')
                goods_titles = element.select('div.p-name.p-name-type-2 a em')
                goods_prices = element.select('div.p-price strong i')
                goods_comments = element.select('div.p-commit strong a')

                total_num += 1
                # 筛选
                if (len(shop_elements) != 0):
                    if filter_by_shop_name(shop_elements[0].text):
                        continue
                if len(shop_elements) == 0:
                    continue
                
                # 筛选掉评论数不足200的商品
                if(len(goods_comments) != 0):
                    if convert_string_to_number(goods_comments[0].text) < 200:
                        continue
                record_num += 1
                
                # 下一行
                last_column = 0
                
                # 序号
                try:
                    last_column+=1
                    ordinal = last_row + index
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['ordinal'] = ordinal
                
                # 电商平台
                try:
                    last_column+=1
                    platform_name = '京东'
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['platform'] = platform_name
                
                # 关键词
                try:
                    last_column+=1
                    search_keyword = keyword
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['keyword'] = search_keyword
                
                # 店铺名称
                try:
                    last_column+=1
                    shop_name = shop_elements[0].text if len(shop_elements) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['shopname'] = shop_name
                
                # 店铺网址
                try:
                    last_column+=1
                    shop_link = 'https:' + shop_elements[0].get('href') if len(shop_elements) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['shoplink'] = shop_link
                
                # 店铺经营主体信息
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['body'] = ''

                # 商品图片
                try:
                    last_column+=1
                    if(len(goods_elements) != 0):
                        goods_img_urls = goods_elements[0].select('img')
                        if (len(goods_img_urls) != 0):
                            goods_img_url = goods_img_urls[0].get('src')
                            if goods_img_url:
                                goods_img_url = 'https:' + goods_img_url
                                if goods_img_url.endswith('.avif'):
                                    goods_img_url = goods_img_url[:-5]
                            else:
                                goods_img_url = 'https:' + (goods_elements[0].select('img')[0].get('data-lazy-img'))
                                if goods_img_url.endswith('.avif'):
                                    goods_img_url = goods_img_url[:-5]
                        else:
                            goods_img_url = ''
                    else:
                        goods_img_url = ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['imgurl'] = goods_img_url
                
                # 商品标题
                try:
                    last_column+=1
                    goods_title = goods_titles[0].text if len(goods_titles) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['title'] = goods_title

                # 商品品牌
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['brand'] = ''
                
                # 商品链接
                try:
                    last_column+=1
                    goods_link = 'https:' + goods_elements[0].get('href') if len(goods_elements) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['goodslink'] = goods_link
                
                # 单价
                try:
                    last_column+=1
                    goods_price = goods_prices[0].text if len(goods_prices) != 0 else '0'
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['price'] = goods_price
                
                # 销售量
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['amount'] = ''
                
                # 商品评论数
                try:
                    last_column+=1
                    goods_comment = (goods_comments[0].text if goods_comments[0].text else '0') if len(goods_comments) != 0 else '0'
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['comment'] = goods_comment
                
                # 销售额
                try:
                    last_column+=1
                    if (len(goods_prices) != 0 and len(goods_comments) != 0):
                        goods_price = is_float(goods_prices[0].text) and float(goods_prices[0].text) or 0
                        goods_comment = convert_string_to_number(goods_comments[0].text)
                        goods_sales = goods_price * goods_comment
                    else:
                        goods_sales = 0
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['sales'] = goods_sales

                item_page.append(item_row)
        except Exception as e:
            print(e)
            print('与现有浏览器连接断开')

        return [total_num, record_num, item_page]

    def file_prepare(self, keyword):
        # 获取当前时间并格式化时间字符串，将其作为文件名一部分
        current_time = datetime.datetime.now()
        time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        # 构建文件名，格式为：平台名称+关键词+时间
        file_path = f"../data/jd/京东_{keyword}_{time_string}.xlsx"
        # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
        self.total_num = 0
        self.record_num = 0

        # 创建一个新的excel用于记录数据操作并设置好表头为headers的内容，并将得到的新表初步保存一下
        workbook = Workbook()
        sheet = workbook.active
        headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
        sheet.append(headers)
        workbook.save(file_path)
        return [file_path, headers]
    
    def get_max_page(self, soup):
        # 通过bs4来获取该关键词的搜索得到的商品最大页数，纠正end_page
        try:
            max_page = 1
            end_page = self.end_page
            elements = soup.select('span.p-skip em b')
            if len(elements):
                max_page = int(elements[0].text)*2
            print(f"期望页数：{end_page}，最大页数：{max_page}")
            if(end_page > max_page):
                end_page = max_page
        except Exception as e:
            print(e)
            print('获取最大页数时出错')
        finally:
            return end_page
    
    def verify_monitor(self, driver):
        html = driver.execute_script("return document.documentElement.outerHTML")
        soup = BeautifulSoup(html, 'html.parser')
        verify = soup.select('div.verifyBtn')
        while len(verify) != 0:
            firefox_window = gw.getWindowsWithTitle("Mozilla Firefox")[0]
            firefox_window.minimize()
            firefox_window.maximize()
            firefox_window.activate()
            print('遇到验证码')
            time.sleep(8)
            html = driver.execute_script(
                "return document.documentElement.outerHTML")
            soup = BeautifulSoup(html, "html.parser")
            verify = soup.select('div.verifyBtn')
            
# 店铺名称筛选
def filter_by_shop_name(shopName):
    keywords = ['华为京东自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

# 商品标题筛选
def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

# 商品评论数筛选
def filter_by_goods_commit(goodsCommit):
    if not goodsCommit:
        return True
    if goodsCommit.endswith('万+'):
        return False
    elif goodsCommit.endswith('+'):
        if int(goodsCommit[:-1])>=200:
            return False
        else:
            return True
    else:
        return True

# 字符串转数字
def convert_string_to_number(string):
    if not string:
        return 0
    if string == '':
        return 0
    if string.endswith('万+'):
        number = int(string[:-2]) * 10000
    elif string.endswith('+'):
        number = int(string[:-1])
    else:
        number = int(string)
    return number

def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False
    
