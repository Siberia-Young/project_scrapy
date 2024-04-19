from scrapy_selenium import SeleniumRequest
from openpyxl import Workbook
import scrapy
import datetime
import json
import os
from openpyxl import load_workbook
import glob

from project_scrapy.modules.tools.tip import tip

from project_scrapy.items import Item

class PDDSpider(scrapy.Spider):
    name = "pdd"
    def start_requests(self):
        tip(self.name)
        self.min_delay = 3  # 最小延迟时间（单位：秒）
        self.max_delay = 5  # 最大延迟时间（单位：秒）

        # 读取parameter.json的参数
        with open('parameter.json', encoding='utf-8') as file:
            parameter = json.load(file)
            self.keywords = parameter['keywords']
            self.relatedwords = parameter['relatedwords']
            self.graphicwords = parameter['graphicwords']
            self.start_page = parameter['start_page']
            self.end_page = parameter['end_page']
        
        self.process_status = {}
        for keyword in self.keywords:
            self.process_status[keyword] = False

        url = 'https://pifa.pinduoduo.com'
        yield SeleniumRequest(url=url, callback=self.parse, wait_time=2)
        
    def parse(self, response):
        folder_path = "../data/pdd/json"
        json_files = glob.glob(os.path.join(folder_path, "*.json"))
        keywords = [os.path.basename(file).replace('.json','') for file in json_files]
        for keyword in keywords:
            [file_path, headers] = self.file_prepare(keyword)
            # 读取JSON文件
            elements = []
            with open(os.path.join(folder_path, f'{keyword}.json').replace(os.sep, '/'), 'r', encoding='utf-8') as file:
                elements = json.load(file)

            total_num = 0
            record_num = 0
            try:
                [single_total_num, single_record_num, item_page] = self.parse_page(elements, headers, file_path, keyword)
                total_num += single_total_num
                record_num += single_record_num

                item = Item()
                item['category'] = 'PDD_spider'
                item['filepath'] = file_path
                item['keyword'] = keyword
                item['page'] = 1
                item['goods'] = item_page
                yield item
            except Exception as e:
                print(e)
                print('记录每页数据时发生错误')
            except KeyboardInterrupt:
                print('用户主动中断爬虫')
            finally:
                # 重命名文件
                new_file_path = file_path.replace('.xlsx',f'_({record_num} of {total_num}).xlsx')
                try:
                    os.rename(file_path, new_file_path)
                    print(f"已将文件 {file_path} 重命名为 {new_file_path}")
                except Exception as e:
                    print(e)
                    print(f"重命名文件 {file_path} 失败")
                finally:
                    print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")
                    self.process_status[keyword] = True
                    print(f'关键词爬取完毕：{keyword}')
                    
                    item = Item()
                    item['category'] = 'PDD_process'
                    item['end'] = True
                    item['keywords'] = self.keywords
                    item['relatedwords'] = self.relatedwords
                    item['graphicwords'] = self.graphicwords
                    for value in self.process_status.values():
                        if not value:
                            item['end'] = False
                    yield item

    def parse_page(self, elements, headers, file_path, keyword):
        # 加载之前创建的excel表
        workbook = load_workbook(file_path)
        sheet = workbook.active
        last_row = sheet.max_row

        # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
        total_num = 0
        record_num = 0
        item_page = []

        try:
            # 逐个解析提取列表中商品信息的各部分数据
            for i in range(len(elements)):
                for element in elements[i]:
                    item_row = {}

                    total_num += 1
                    record_num += 1
                    
                    # 下一行
                    last_row+=1
                    last_column = 0
                    
                    # 序号
                    try:
                        last_column+=1
                        ordinal = last_row-1
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['ordinal'] = ordinal
                    
                    # 电商平台
                    try:
                        last_column+=1
                        platform_name = '拼多多-批发'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['platform'] = platform_name
                    
                    # 关键词
                    try:
                        last_column+=1
                        search_keyword = keyword
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['keyword'] = search_keyword
                    
                    # 店铺名称
                    try:
                        last_column+=1
                        shop_name = 'mallName' in element.keys() and element['mallName'] or '暂无店铺名称'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['shopname'] = shop_name
                    
                    # 店铺网址
                    try:
                        last_column+=1
                        shop_link = 'mallIdEncrypt' in element.keys() and ('https://pifa.pinduoduo.com/mall?mid='+element['mallIdEncrypt']) or '暂无店铺链接'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['shoplink'] = shop_link
                    
                    # 店铺经营主体信息
                    try:
                        last_column+=1
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['body'] = ''

                    # 商品图片
                    try:
                        last_column+=1
                        goods_img_url = 'goodsImgUrl' in element.keys() and element['goodsImgUrl'] or '暂无商品图片'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['imgurl'] = goods_img_url
                    
                    # 商品标题
                    try:
                        last_column+=1
                        goods_title = 'goodsName' in element.keys() and element['goodsName'] or '暂无商品标题'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['title'] = goods_title

                    # 商品品牌
                    try:
                        last_column+=1
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['brand'] = ''
                    
                    # 商品链接
                    try:
                        last_column+=1
                        goods_link = 'goodsId' in element.keys() and 'https://pifa.pinduoduo.com/goods/detail/?gid='+str(element['goodsId']) or '暂无商品链接'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['goodslink'] = goods_link
                    
                    # 单价
                    try:
                        last_column+=1
                        goods_price = 'goodsWholeSalePrice' in element.keys() and (element['goodsWholeSalePrice']/100) or '暂无单价'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['price'] = goods_price
                    
                    # 销售量
                    try:
                        last_column+=1
                        goods_num = 'salesTipAmount' in element.keys() and element['salesTipAmount'] or '0'
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['amount'] = goods_num
                    
                    # 商品评论数
                    try:
                        last_column+=1
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['comment'] = ''
                    
                    # 销售额
                    try:
                        last_column+=1
                        goods_sales = goods_price * convert_string_to_number(goods_num)
                    except:
                        print(f'记录“{headers[last_column-1]}”时出错')
                        return
                    finally:
                        item_row['sales'] = goods_sales

                    item_page.append(item_row)
        except Exception as e:
            print(e)
            print('与现有浏览器连接断开')

        return [total_num, record_num, item_page]
    
    def file_prepare(self, keyword):
        # 获取当前时间并格式化时间字符串，将其作为文件名一部分
        current_time = datetime.datetime.now()
        time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        # 构建文件名，格式为：平台名称+关键词+时间
        file_path = f"../data/pdd/拼多多批发_{keyword}_{time_string}.xlsx"
        # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
        self.total_num = 0
        self.record_num = 0

        # 创建一个新的excel用于记录数据操作并设置好表头为headers的内容，并将得到的新表初步保存一下
        workbook = Workbook()
        sheet = workbook.active
        headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
        sheet.append(headers)
        workbook.save(file_path)
        return [file_path, headers]



def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = float(string[:-2]) * 10000
    elif string.endswith('万'):
        number = float(string[:-1]) * 10000
    elif string.endswith('+'):
        number = float(string[:-1])
    else:
        number = float(string)
    return number