from scrapy_selenium import SeleniumRequest
from openpyxl import Workbook
from bs4 import BeautifulSoup
import urllib.parse
import scrapy
import datetime
import random
import json
import time
import os
from openpyxl import load_workbook
import re
import math

from project_scrapy.modules.tools.tip import tip

from project_scrapy.items import Item

class O1688Spider(scrapy.Spider):
    name = "1688"
    def start_requests(self):
        tip(self.name)
        self.min_delay = 3  # 最小延迟时间（单位：秒）
        self.max_delay = 5  # 最大延迟时间（单位：秒）

        # 读取parameter.json的参数
        with open('parameter.json', encoding='utf-8') as file:
            parameter = json.load(file)
            self.keywords = parameter['keywords']
            self.relatedwords = parameter['relatedwords']
            self.graphicwords = parameter['graphicwords']
            self.start_page = parameter['start_page']
            self.end_page = parameter['end_page']
        
        self.process_status = {}
        for keyword in self.keywords:
            self.process_status[keyword] = False

        url = 'https://s.1688.com'
        yield SeleniumRequest(url=url, callback=self.parse, wait_time=2)
        
    def parse(self, response):
        driver = response.webdriver
        for keyword in  self.keywords:
            [file_path, headers] = self.file_prepare(keyword)
            
            # 通过构造url来访问关键词的搜索结果，等待部分时间加载页面，用bs4来解析页面
            driver.get(f"https://s.1688.com/selloffer/offer_search.htm?keywords={urllib.parse.quote(keyword, encoding='GBK')}&sortType=va_rmdarkgmv30")
            time.sleep(2)

            html = driver.execute_script("return document.documentElement.outerHTML")
            # 创建 Beautiful Soup 对象
            soup = BeautifulSoup(html, "html.parser")

            start_page = self.start_page
            end_page = self.get_max_page(soup)
            total_num = 0
            record_num = 0
            try:
                for page in range(start_page, end_page+1):
                    html = driver.execute_script("return document.documentElement.outerHTML")
                    # 创建 Beautiful Soup 对象
                    soup = BeautifulSoup(html, "html.parser")
                    end_page = self.get_max_page(soup)
                    if(page > end_page):
                        break
                    [single_total_num, single_record_num, item_page] = self.parse_page(driver, page, headers, file_path, keyword)
                    total_num += single_total_num
                    record_num += single_record_num
                    if single_record_num == 0:
                        break
                    item = Item()
                    item['category'] = '1688_spider'
                    item['filepath'] = file_path
                    item['keyword'] = keyword
                    item['page'] = page
                    item['goods'] = item_page
                    yield item
            except Exception as e:
                print(e)
                print('记录每页数据时发生错误')
            except KeyboardInterrupt:
                print('用户主动中断爬虫')
            finally:
                # 重命名文件
                new_file_path = file_path.replace('.xlsx',f'_({record_num} of {total_num}).xlsx')
                try:
                    os.rename(file_path, new_file_path)
                    print(f"已将文件 {file_path} 重命名为 {new_file_path}")
                except Exception as e:
                    print(e)
                    print(f"重命名文件 {file_path} 失败")
                finally:
                    print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")
                    self.process_status[keyword] = True
                    print(f'关键词爬取完毕：{keyword}')
                    
                    item = Item()
                    item['category'] = '1688_process'
                    item['end'] = True
                    item['keywords'] = self.keywords
                    item['relatedwords'] = self.relatedwords
                    item['graphicwords'] = self.graphicwords
                    for value in self.process_status.values():
                        if not value:
                            item['end'] = False
                    yield item

    def parse_page(self, driver, page, headers, file_path, keyword):
        # 加载之前创建的excel表
        workbook = load_workbook(file_path)
        sheet = workbook.active
        last_row = sheet.max_row

        # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
        total_num = 0
        record_num = 0

        # 设置适当的时间间隔以避免触发反爬虫
        time.sleep(random.uniform(self.min_delay, self.max_delay))

        self.simulate_page(driver, page)

        self.simulate_slide(driver)
        
        # 通过bs4得到整个页面解析得来的单个商品元素构成的列表
        html = driver.execute_script("return document.documentElement.outerHTML")
        soup = BeautifulSoup(html, "html.parser") 
        # 使用 select 方法查找指定的元素
        elements = soup.select('div.space-offer-card-box')
        
        item_page = []
        try:
            # 逐个解析提取列表中商品信息的各部分数据
            for (index, element) in enumerate(elements):
                item_row = {}

                shop_elements = element.find_all('div', class_='company-name', title=True)
                goods_elements = element.select('div.mojar-element-image a')
                goods_titles = element.select('div.mojar-element-title a div.title')
                goods_prices = element.select('div.showPricec div.price')
                goods_sales = element.select('div.sale div.count')
                
                total_num += 1
                
                # 筛选
                if (len(shop_elements) != 0):
                    if filter_by_shop_name(shop_elements[0].text):
                        continue
                if len(shop_elements) == 0:
                    continue
                if(len(goods_prices) == 0):
                    continue
                # 筛选掉销售额为0的商品
                if(len(goods_sales) != 0):
                    if goods_sales[0].text == '':
                        continue
                if(len(goods_prices) != 0):
                    if not is_float(goods_prices[0].text):
                        continue

                record_num += 1
                
                # 下一行
                last_column = 0
                
                # 序号
                try:
                    last_column+=1
                    ordinal = last_row + index
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['ordinal'] = ordinal
                
                # 电商平台
                try:
                    last_column+=1
                    platform_name = '1688'
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['platform'] = platform_name
                
                # 关键词
                try:
                    last_column+=1
                    search_keyword = keyword
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['keyword'] = search_keyword
                
                # 店铺名称
                try:
                    last_column+=1
                    shop_name = shop_elements[0].get('title') if len(shop_elements) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['shopname'] = shop_name
                
                # 店铺网址
                try:
                    last_column+=1
                    shop_link = (shop_elements[0].select('a')[0].get('href') if len(shop_elements[0].select('a')) != 0 else '') if len(shop_elements) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['shoplink'] = shop_link
                
                # 店铺经营主体信息
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['body'] = ''

                # 商品图片
                try:
                    last_column+=1
                    goods_img_url = ''
                    if(len(goods_elements) != 0):
                        goods_img_urls = goods_elements[0].select('div.img')
                        if (len(goods_img_urls) != 0):
                            goods_img_url_style = goods_img_urls[0].get('style')
                            if goods_img_url_style:
                                goods_img_url = re.search(r"url\(.*?[\'\"](.*?)['\"]\)", goods_img_url_style)
                                if goods_img_url:
                                    goods_img_url = goods_img_url.group(1)
                                    goods_img_url = goods_img_url.replace("?_=2020","")
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['imgurl'] = goods_img_url
                
                # 商品标题
                try:
                    last_column+=1
                    goods_title = goods_titles[0].text if len(goods_titles) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['title'] = goods_title

                # 商品品牌
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['brand'] = ''
                
                # 商品链接
                try:
                    last_column+=1
                    goods_link = goods_elements[0].get('href') if len(goods_elements) != 0 else ''
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['goodslink'] = goods_link
                
                # 单价
                try:
                    last_column+=1
                    goods_price = goods_prices[0].text if len(goods_prices) != 0 else '0'
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['price'] = goods_price
                
                # 销售量
                try:
                    last_column+=1
                    goods_num = 0
                    if len(goods_sales) != 0:
                        goods_sales_text = goods_sales[0].text
                        if (goods_sales_text.startswith('成交') and goods_sales_text.endswith('元')):
                            goods_sales_text = goods_sales_text.replace('成交','').replace('元','')
                        else:
                            goods_sales_text = '0'
                        goods_num = convert_string_to_number(goods_sales_text)/float(goods_price)
                        goods_num = math.ceil(goods_num)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['amount'] = goods_num
                
                # 商品评论数
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['comment'] = ''
                
                # 销售额
                try:
                    last_column+=1
                    sales = 0
                    if (len(goods_sales) != 0):
                        goods_sales_text = goods_sales[0].text
                        if (goods_sales_text.startswith('成交') and goods_sales_text.endswith('元')):
                            sales = convert_string_to_number(goods_sales_text.replace('成交','').replace('元',''))
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    return
                finally:
                    item_row['sales'] = sales

                item_page.append(item_row)
        except Exception as e:
            print(e)
            print('与现有浏览器连接断开')

        return [total_num, record_num, item_page]
    
    def file_prepare(self, keyword):
        # 获取当前时间并格式化时间字符串，将其作为文件名一部分
        current_time = datetime.datetime.now()
        time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        # 构建文件名，格式为：平台名称+关键词+时间
        file_path = f"../data/1688/1688_{keyword}_{time_string}.xlsx"
        # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
        self.total_num = 0
        self.record_num = 0

        # 创建一个新的excel用于记录数据操作并设置好表头为headers的内容，并将得到的新表初步保存一下
        workbook = Workbook()
        sheet = workbook.active
        headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
        sheet.append(headers)
        workbook.save(file_path)
        return [file_path, headers]
    
    def get_max_page(self, soup):
        try:
            max_page = 1
            end_page = self.end_page
            elements = soup.select('em.fui-paging-num')
            if len(elements):
                max_page = int(elements[0].text)
            print(f"期望页数：{end_page}，最大页数：{max_page}")
            if(end_page > max_page):
                end_page = max_page
        except Exception as e:
            print(e)
            print('获取最大页数时出错')
        finally:
            return end_page

    def simulate_page(self, driver, page):
        try:
            if(page!=self.start_page):
                button = driver.find_element("xpath", "//a[contains(@class, 'fui-next')]")
                driver.execute_script("document.querySelector('a.fui-next').style.position = 'relative';")
                driver.execute_script("document.querySelector('a.fui-next').style.zIndex = '99999';")
                button.click()
                time.sleep(2)
        except:
            print('翻页时出错')
    
    def simulate_slide(self, driver):
        try:
            # 缓慢下拉页面
            scroll_height = driver.execute_script("return document.body.scrollHeight;")
            current_height = 0
            scroll_speed = 500  # 每次下拉的距离
            while current_height < scroll_height:
                driver.execute_script(f"window.scrollTo(0, {current_height});")
                current_height += scroll_speed
                time.sleep(0.3)  # 等待一段时间，模拟缓慢下拉的效果
                scroll_height = driver.execute_script("return document.body.scrollHeight;")
        except:
            print('下拉获取页面信息时发生错误')



# 店铺名称筛选
def filter_by_shop_name(shopName):
    keywords = ['华为京东自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

# 商品标题筛选
def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

# 商品评论数筛选
def filter_by_goods_commit(goodsCommit):
    if not goodsCommit:
        return True
    if goodsCommit.endswith('万+'):
        return False
    elif goodsCommit.endswith('+'):
        if int(goodsCommit[:-1])>=200:
            return False
        else:
            return True
    else:
        return True

# 字符串转数字
def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = int(string[:-2]) * 10000
    elif string.endswith('+'):
        number = int(string[:-1])
    else:
        number = int(string)
    return number

def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False