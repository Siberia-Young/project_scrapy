# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
import os
from openpyxl import load_workbook
from project_scrapy.modules.tools.file_preprocess import file_preprocess
from project_scrapy.modules.tools.file_postprocess import file_postprocess
from project_scrapy.modules.tools.file_select import file_select
from project_scrapy.modules.tools.file_package import file_package
from project_scrapy.modules.tools.merge import merge
from project_scrapy.modules.code_2_filter_by_repeat import filter_by_repeat
from project_scrapy.modules.code_3_filter_by_whitelist import filter_by_whitelist
from project_scrapy.modules.code_4_crawl_and_save_product_images import crawl_and_save_product_images
from project_scrapy.modules.code_5_extract_image_text import extract_image_text
from project_scrapy.modules.code_6_filter_by_image_text import filter_by_image_text
from project_scrapy.modules.code_8_classify_and_sort import classify_and_sort
from project_scrapy.modules.code_9_filter_by_sales import filter_by_sales
from project_scrapy.modules.code_10_filter_by_detailed_data import filter_by_detailed_data
from project_scrapy.modules.code_11_cell_style_adjustments import cell_style_adjustments

keys = ['ordinal', 'platform', 'keyword', 'shopname', 'shoplink', 'body', 'imgurl', 'title', 'brand', 'goodslink', 'price', 'amount', 'comment', 'sales']
category_list = ['JD_spider', 'TB_spider', 'PDD_spider', '1688_spider']

class SpiderPipeline:
    def process_item(self, item, spider):
        if item['category'] in category_list:
            # 加载之前创建的excel表
            workbook = load_workbook(item['filepath'])
            sheet = workbook.active
            last_row = sheet.max_row
            for index, goods in enumerate(item['goods']):
                for key, value in goods.items():
                    sheet.cell(row=last_row+index+1, column=keys.index(key)+1, value=value)
            print(f"正在记录关键词：{item['keyword']} 的第 {str(item['page'])} 页")
            workbook.save(item['filepath'])
            print(f"已保存第 {str(item['page'])} 页数据到 {item['filepath']}")
        return item

class JDProcessPipeline:
    def process_item(self, item, spider):
        [platform_name, source_folder, destination_folder, outcome_folder, final_folder, target_list] = parse_folder('jd')
        if item['category'] == 'JD_process_1':
            if item['end']:
                file_preprocess(source_folder, destination_folder)
                merge(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'), destination_folder)
                my_print(2)
                filter_by_repeat(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'))
                my_print(3)
                filter_by_whitelist(platform_name, item['relatedwords'][0], item['relatedwords'][1], item['relatedwords'][2], os.path.join(destination_folder, "merge_2.xlsx").replace(os.sep, '/'))
        elif item['category'] == 'JD_process_2':
            if item['end']:
                my_print(8)
                classify_and_sort(platform_name, os.path.join(destination_folder, "merge_2_3.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8.xlsx").replace(os.sep, '/'))
        elif item['category'] == 'JD_process_3':
            if item['end']:
                my_print(10)
                filter_by_detailed_data(os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8_9_10.xlsx").replace(os.sep, '/'))
                my_print(4)
                crawl_and_save_product_images(platform_name, os.path.join(destination_folder, "merge_2_3_8_9_10_9.xlsx").replace(os.sep, '/'))
                my_print(5)
                extract_image_text(os.path.join(destination_folder, "merge_2_3_8_9_10_9.xlsx").replace(os.sep, '/'))
                my_print(6)
                filter_by_image_text(item['graphicwords'], os.path.join(destination_folder, "merge_2_3_8_9_10_9.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8_9_10_9_6.xlsx").replace(os.sep, '/'))
                file_select(destination_folder, outcome_folder, target_list)
                my_print(11)
                cell_style_adjustments(outcome_folder)
                file_package(platform_name, source_folder, outcome_folder, final_folder)
                file_postprocess(source_folder)
        return item

class TBProcessPipeline:
    def process_item(self, item, spider):
        [platform_name, source_folder, destination_folder, outcome_folder, final_folder, target_list] = parse_folder('tb')
        if item['category'] == 'TB_process':
            if item['end']:
                file_preprocess(source_folder, destination_folder)
                merge(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'), destination_folder)
                my_print(2)
                filter_by_repeat(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'))
                my_print(3)
                filter_by_whitelist(platform_name, item['relatedwords'][0], item['relatedwords'][1], item['relatedwords'][2], os.path.join(destination_folder, "merge_2.xlsx").replace(os.sep, '/'))
                my_print(8)
                classify_and_sort(platform_name, os.path.join(destination_folder, "merge_2_3.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8.xlsx").replace(os.sep, '/'))
                my_print(4)
                crawl_and_save_product_images(platform_name, os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(5)
                extract_image_text(os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(6)
                filter_by_image_text(item['graphicwords'], os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8_9_6.xlsx").replace(os.sep, '/'))
                file_select(destination_folder, outcome_folder, target_list)
                my_print(11)
                cell_style_adjustments(outcome_folder)
                file_package(platform_name, source_folder, outcome_folder, final_folder)
                file_postprocess(source_folder)
        return item
    
class PDDProcessPipeline:
    def process_item(self, item, spider):
        [platform_name, source_folder, destination_folder, outcome_folder, final_folder, target_list] = parse_folder('pdd')
        if item['category'] == 'PDD_process':
            if item['end']:
                file_preprocess(source_folder, destination_folder)
                merge(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'), destination_folder)
                my_print(2)
                filter_by_repeat(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'))
                my_print(3)
                filter_by_whitelist(platform_name, item['relatedwords'][0], item['relatedwords'][1], item['relatedwords'][2], os.path.join(destination_folder, "merge_2.xlsx").replace(os.sep, '/'))
                my_print(8)
                classify_and_sort(platform_name, os.path.join(destination_folder, "merge_2_3.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8.xlsx").replace(os.sep, '/'))
                my_print(4)
                crawl_and_save_product_images(platform_name, os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(5)
                extract_image_text(os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(6)
                filter_by_image_text(item['graphicwords'], os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8_9_6.xlsx").replace(os.sep, '/'))
                file_select(destination_folder, outcome_folder, target_list)
                my_print(11)
                cell_style_adjustments(outcome_folder)
                file_package(platform_name, source_folder, outcome_folder, final_folder)
                file_postprocess(source_folder)
        return item

class O1688ProcessPipeline:
    def process_item(self, item, spider):
        [platform_name, source_folder, destination_folder, outcome_folder, final_folder, target_list] = parse_folder('1688')
        if item['category'] == '1688_process':
            if item['end']:
                file_preprocess(source_folder, destination_folder)
                merge(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'), destination_folder)
                my_print(2)
                filter_by_repeat(os.path.join(destination_folder, "merge.xlsx").replace(os.sep, '/'))
                my_print(3)
                filter_by_whitelist(platform_name, item['relatedwords'][0], item['relatedwords'][1], item['relatedwords'][2], os.path.join(destination_folder, "merge_2.xlsx").replace(os.sep, '/'))
                my_print(8)
                classify_and_sort(platform_name, os.path.join(destination_folder, "merge_2_3.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8.xlsx").replace(os.sep, '/'))
                my_print(4)
                crawl_and_save_product_images(platform_name, os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(5)
                extract_image_text(os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(6)
                filter_by_image_text(item['graphicwords'], os.path.join(destination_folder, "merge_2_3_8_9.xlsx").replace(os.sep, '/'))
                my_print(9)
                filter_by_sales(platform_name, os.path.join(destination_folder, "merge_2_3_8_9_6.xlsx").replace(os.sep, '/'))
                file_select(destination_folder, outcome_folder, target_list)
                my_print(11)
                cell_style_adjustments(outcome_folder)
                file_package(platform_name, source_folder, outcome_folder, final_folder)
                file_postprocess(source_folder)
        return item
    
def my_print(num):
    str1 = num < 10 and '0' + str(num) or str(num)
    print(f'\n----------------{str1}----------------')

def parse_folder(platform_name):
    source_folder = f"../data/{platform_name}"
    destination_folder = f"../data/{platform_name}/merge"
    outcome_folder = f"../data/{platform_name}/merge/outcome"
    final_folder = f"../data/z_submit"
    if platform_name == 'jd':
        target_list = ['merge.xlsx', 'merge_2.xlsx', 'merge_2_3.xlsx', 'merge_2_3_8_9(副本).xlsx', 'merge_2_3_8_9_10_9_6.xlsx', 'merge_2_3_8_9_10_9(副本).xlsx']
    elif platform_name == 'tb' or platform_name == 'pdd' or platform_name == '1688':
        target_list = ['merge.xlsx', 'merge_2.xlsx', 'merge_2_3.xlsx', 'merge_2_3_8_9(副本).xlsx', 'merge_2_3_8_9_6_9.xlsx', 'merge_2_3_8_9(副本).xlsx']
    else:
        target_list = []
        print('没有此平台')
    return [platform_name, source_folder, destination_folder, outcome_folder, final_folder, target_list]