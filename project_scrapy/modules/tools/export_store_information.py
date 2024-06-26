from openpyxl import load_workbook
import time
import os
import datetime
import json

def export_store_information(platform_name, file_path):
    base_file_path = f'data_files/{platform_name}/shop_body_info/shop_body_info.json'
    current_time = datetime.datetime.now()
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    new_file_path = base_file_path.replace('.json',f'_{time_string}.json')
    # 打开需读取的excel表
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 读取现有店铺信息
    try:
        print(f'\n正在读取现有店铺信息')
        with open(base_file_path, encoding='utf-8') as file:
            dict = json.load(file)
        print(f'\n读取现有信息：{len(dict)}')
    except Exception as e:
        print(e)
        print('读取现有店铺信息时出错')

    # 筛选需要新记录的店铺信息
    try:  
        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在筛选需要新记录的店铺信息')
        for row in range(start_row, end_row + 1):
            shop_name = sheet.cell(row=row, column=4).value
            shop_link = sheet.cell(row=row, column=5).value
            shop_body = sheet.cell(row=row, column=6).value
            if shop_link not in dict and shop_body is not None:
                dict[shop_link] = [shop_name,shop_body]
    except Exception as e:
        print(e)
        print('筛选需要新记录的店铺信息时出错')

    # 记录到新的json
    try:
        print(f'\n正在记录到新的json')
        with open(new_file_path, 'w', encoding='utf-8') as file:
            json.dump(dict, file, indent=4, ensure_ascii=False)
        print(f'\n成功更新json：{len(dict)}')
    except Exception as e:
        print(e)
        print('记录到新的json时出错')

    # 修改文件名
    try:
        temp_file_path = "/".join(base_file_path.split("/")[:-1]) + '/temp.json'
        os.rename(base_file_path, temp_file_path)
        os.rename(new_file_path, base_file_path)
        os.rename(temp_file_path, new_file_path)
    except Exception as e:
        print(e)
        print('修改文件名时出错')