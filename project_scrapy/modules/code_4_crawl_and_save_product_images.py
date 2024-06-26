from openpyxl import load_workbook
import time
import requests
import os
import json

num = 4
headers = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
}

def crawl_and_save_product_images(platform_name, file_path):
    folder_path = "/".join(file_path.split("/")[:-1]) + '/images'
    # 打开需读取的excel表
    workbook = load_workbook(file_path)
    sheet = workbook.active

    if platform_name == 'jd' or platform_name == 'pdd':
        # 处理文件夹不存在的情况
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        
        # 下载图片
        try:
            fail_list = []
            start_row = 2
            end_row = sheet.max_row

            total = end_row - start_row + 1
            current = 0
            start_time = time.time()
            time.sleep(1)
            print(f'\n正在下载图片')
            for row in range(start_row, end_row + 1):
                current += 1
                res = (total - current) / (current /
                                        ((time.time() - start_time) / 60))
                print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
                value = sheet.cell(row=row, column=7).value
                response = requests.get(value)
                if response.status_code == 200:
                    image_path = os.path.join(
                        folder_path, f'{row}.{value.split(".")[-1]}')
                    with open(image_path, 'wb') as file:
                        file.write(response.content)
                else:
                    fail_list.append(row)
        except Exception as e:
            print(e)
            print("关闭VPN")
        finally:
            end_time = time.time()
            duration = end_time - start_time
            print(f"下载耗时：{(duration/60):.2f} min")
            print(f"目标数量：{total} 条")
            print(f"已下载数量：{current} 条")
            unit = current / (duration / 60)
            print(f"每分钟下载数量：{unit:.2f} 条")

        print(f'下载失败：{fail_list}')


    elif platform_name == 'tb' or platform_name == '1688':
        init = True
        # 处理文件夹不存在的情况
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        else:
            init = False

        # 从fail_list.json中获取数据
        if init:
            numbers = list(range(2, sheet.max_row+1))
            with open(folder_path+'/1.json', 'w') as file:
                json.dump(numbers, file)

        with open(folder_path+'/1.json', 'r') as file:
            my_list = json.load(file)

        # 下载图片
        try:
            comeon = True
            while(comeon):
                ip = input('请输入ip和端口：')
                # ip = '111.177.63.86:8888'
                fail_list = []
                start_row = 2
                end_row = sheet.max_row

                total = end_row - start_row + 1
                current = 0
                start_time = time.time()
                time.sleep(1)
                print(f'\n正在下载图片')
                total = len(my_list)
                for row in my_list:
                    current+=1
                    res = (total - current) / (current / ((time.time() - start_time) / 60))
                    print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
                    value = sheet.cell(row=row, column=7).value


                    response = requests.get(value,proxies={'http':ip,'https':ip},headers=headers)
                    # response = requests.get(value)
                    if response.status_code == 200:
                        image_path = os.path.join(folder_path, f'{row}.{value.split(".")[-1]}')
                        with open(image_path, 'wb') as file:
                            file.write(response.content)
                    else:
                        fail_list.append(row)
                print('失败列表：',fail_list)
                comeon = input('是否再次爬取图片？(y/n)') == 'y'
        except Exception as e:
            print(e)
            print("关闭VPN")
        finally:
            if current == total:
                with open(folder_path+'/1.json', 'w') as file:
                    json.dump(fail_list, file)
            else:
                with open(folder_path+'/1.json', 'w') as file:
                    json.dump(fail_list + my_list[current-1:], file)
            end_time = time.time()
            duration = end_time - start_time
            print(f"下载耗时：{(duration/60):.2f} min")
            print(f"目标数量：{total} 条")
            print(f"已下载数量：{current} 条")
            unit = current / (duration / 60)
            print(f"每分钟下载数量：{unit:.2f} 条")

        print(f'下载失败：{fail_list}{len(fail_list)}')