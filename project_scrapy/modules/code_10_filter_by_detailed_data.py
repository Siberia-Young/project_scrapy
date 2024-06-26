# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import time
from openpyxl.utils.cell import get_column_letter
import re

num = 10

def filter_by_detailed_data(file_path):
    new_file_path = file_path.replace('.xlsx','_') + str(num) + '.xlsx'
    # 打开需读取的excel表
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 新建excel表
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # 处理表头
    print(f'\n正在处理表头')
    first_row = sheet[1]
    for cell in first_row:
        new_sheet[cell.coordinate].value = cell.value

    # 通过详细信息筛选
    try:
        record_list = []
        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在通过详细信息筛选')
        def check_keywords(text):
            if text is None:
                return True
            keywords = ['华为']
            pattern = '|'.join(keywords)
            match = re.search(pattern, text, flags=re.IGNORECASE)
            return match is not None
        for row in range(start_row, end_row + 1):
            delete = sheet.cell(row=row, column=15).value
            choose = sheet.cell(row=row, column=17).value
            parameter = sheet.cell(row=row, column=18).value
            differ = sheet.cell(row=row, column=19).value
            # if delete != 'delete' and (check_keywords(choose) or check_keywords(parameter) or differ == 'different'):
            #     record_list.append(row)
            if delete != 'delete':
                record_list.append(row)
    except Exception as e:
        print(e)
        print('通过详细信息筛选时出错')

    # 记录数据到新表
    try:
        start_row = 2
        end_row = sheet.max_row

        total = len(record_list)
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在记录数据到新表')
        for row in record_list:
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            for cell in sheet[row][:14]:
                new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
    except Exception as e:
        print(e)
        print('记录数据到新表时出错')

    # 处理序号
    try:
        start_row = 2
        end_row = new_sheet.max_row

        total = total
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在处理序号')
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            new_sheet.cell(row=row, column=1, value=row-1)
    except Exception as e:
        print(e)
        print('处理序号时出错')

    new_workbook.save(new_file_path)